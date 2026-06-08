import openpyxl
from openpyxl.styles import Font, PatternFill

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "NTB"

# Row 1 Metadata
ws['B1'] = "Thời gian cập nhật"
ws['C1'] = "08/06/2026"

# Row 2 Metadata
ws['B2'] = "Tổng số lượng bưu cục cảnh báo:"
ws['C2'] = 12
ws['C2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws['C2'].font = Font(bold=True, color="FF0000", size=12)

# Row 4 Column Headers
headers = [
    "ngay", "vung_giao", "tinh_giao", "kho_giao_id", "kho_giao_name", 
    "warehouse_type", "BL LM", "BL LM >5 ngay", "%BL LM >5 ngay", 
    "BL KTC", "BL KTC cung tinh %", "%BL KTC cung tinh", "tao_n1", 
    "tao_avg_7ngay", "gtc_n1", "gtc_avg_7ngay", "gtc_max_7ngay", 
    "du_kien_clear_ton", "ly_do_bat_on", "Trạng thái"
]

for col_idx, h in enumerate(headers, start=1):
    ws.cell(row=4, column=col_idx, value=h).font = Font(bold=True)

# 23 rows of test data matching the screenshot
data = [
    # Bất ổn (12 bưu cục)
    ["2026-06-08", "NTB", "Lâm Đồng", 22425000, "Bưu Cục Thôn Phúc Hưng", "BC", 1880, 406, 0.215, 572, 94, 0.164, 300, 405, 376, 320, 381, 6, "2. Tồn LM | 4. Tồn Aging >5 ngày", "Bất ổn"],
    ["2026-06-08", "NTB", "Khánh Hòa", 22704000, "Bưu Cục 05 Lê Hồng Phong", "BC", 2432, 53, 0.022, 925, 162, 0.175, 730, 1081, 860, 757, 913, 3, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 22222000, "Bưu Cục 337 Hùng Vương", "BC", 899, 50, 0.056, 557, 37, 0.066, 262, 378, 406, 297, 406, 3, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Khánh Hòa", 22830000, "Bưu Cục 56 Phan Đình Phùng", "BC", 1844, 100, 0.054, 871, 322, 0.37, 681, 885, 487, 650, 781, 3, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 20942000, "Bưu Cục 1322 Hùng Vương", "BC", 1743, 201, 0.115, 1388, 550, 0.396, 610, 730, 557, 644, 867, 3, "2. Tồn LM | 4. Tồn Aging >5 ngày", "Bất ổn"],
    ["2026-06-08", "NTB", "Đắk Nông", 22242000, "Bưu Cục 55 Tôn Đức Thắng", "BC", 509, 9, 0.018, 297, 122, 0.411, 182, 243, 106, 207, 280, 3, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 21377000, "Bưu Cục TDP Nghĩa Đức", "BC", 1621, 157, 0.097, 1014, 168, 0.166, 584, 822, 728, 749, 900, 2, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 22051000, "Bưu Cục Langbiang", "BC", 829, 38, 0.046, 789, 293, 0.371, 356, 463, 405, 371, 473, 2, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 22759000, "Bưu Cục 231 Thôn 1", "BC", 810, 64, 0.079, 451, 51, 0.113, 234, 389, 348, 393, 467, 2, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Đắk Nông", 22394000, "Bưu Cục Thôn 2", "BC", 648, 60, 0.093, 399, 178, 0.521, 243, 325, 320, 314, 387, 2, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Lâm Đồng", 22312000, "Bưu Cục Số 6 Trương Văn Hoàn", "BC", 1501, 4, 0.003, 1044, 68, 0.065, 654, 918, 683, 749, 868, 2, "2. Tồn LM", "Bất ổn"],
    ["2026-06-08", "NTB", "Khánh Hòa", 21094000, "Kho Giao Hàng Nha Trang", "GXT", 714, 0, 0.000, 470, 198, 0.421, 376, 531, 365, 352, 405, 2, "2. Tồn LM", "Bất ổn"],
    
    # Chuẩn bị nhảy nhóm (8 bưu cục)
    ["2026-06-08", "NTB", "Đắk Nông", 20269000, "Bưu Cục 01 Nguyễn Tất Thành", "BC", 657, 45, 0.068, 520, 215, 0.413, 313, 392, 286, 333, 401, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Lâm Đồng", 22389000, "Bưu Cục Thôn R'Chai 2", "BC", 841, 12, 0.014, 636, 39, 0.061, 350, 482, 367, 445, 589, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Đắk Nông", 22048000, "Bưu Cục Hùng Vương", "BC", 544, 9, 0.017, 409, 121, 0.296, 256, 314, 300, 281, 315, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Khánh Hòa", 22363000, "Bưu Cục Đường 25/4", "BC", 1346, 35, 0.026, 749, 251, 0.335, 532, 859, 751, 713, 823, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Lâm Đồng", 21537000, "Bưu Cục 190/11 Phù Đổng Thiên Vương", "BC", 949, 4, 0.004, 669, 53, 0.079, 402, 625, 483, 514, 660, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Lâm Đồng", 21687000, "Bưu Cục Cao Bá Quát", "BC", 1972, 9, 0.005, 1510, 225, 0.149, 873, 1311, 903, 1092, 1291, 2, "", "Chuẩn bị nhảy nhóm"],
    ["2026-06-08", "NTB", "Khánh Hòa", 20588000, "Bưu Cục 42 Nguyễn Du", "BC", 921, 7, 0.008, 877, 293, 0.334, 521, 664, 646, 504, 646, 2, "", "Chuẩn bị nhảy nhóm"],
    
    # Bình thường (3 bưu cục)
    ["2026-06-08", "NTB", "Ninh Thuận", 21163000, "Kho Giao Hàng Phan Rang", "GXT", 121, 0, 0.000, 250, 57, 0.228, 110, 120, 115, 118, 120, 0, "", "Bình thường"],
    ["2026-06-08", "NTB", "Bình Thuận", 21298000, "Bưu Cục Thị Trấn Liên Hương", "BC", 574, 0, 0.000, 1125, 57, 0.051, 400, 410, 395, 399, 405, 0, "", "Bình thường"],
    ["2026-06-08", "NTB", "Lâm Đồng", 21320000, "Bưu Cục Phú Thủy", "BC", 312, 1, 0.003, 140, 25, 0.178, 150, 155, 148, 151, 153, 0, "", "Bình thường"]
]

for row_idx, row_data in enumerate(data, start=5):
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save("buu_cuc_bat_on.xlsx")
print("New mock buu_cuc_bat_on.xlsx matching columns A to T created successfully.")
