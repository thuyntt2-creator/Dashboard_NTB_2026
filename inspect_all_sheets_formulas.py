import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "inspect_all_sheets_formulas_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for name in wb.sheetnames:
        sheet = wb[name]
        non_empty = 0
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                if sheet.cell(row=r, column=c).value is not None:
                    non_empty += 1
        f.write(f"Sheet '{name}': max_row={sheet.max_row}, max_column={sheet.max_column}, non_empty_cells={non_empty}\n")

print("Done writing inspect_all_sheets_formulas_res.txt")
