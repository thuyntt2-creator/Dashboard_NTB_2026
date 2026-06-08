import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "inspect_formulas_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['Thứ cùng kỳ']
    f.write(f"Thứ cùng kỳ dimensions: max_row={sheet.max_row}, max_column={sheet.max_column}\n")
    
    non_empty_rows = []
    for r in range(1, min(100, sheet.max_row + 1)):
        row_vals = []
        for c in range(1, min(25, sheet.max_column + 1)):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                row_vals.append((c, val))
        if row_vals:
            non_empty_rows.append((r, row_vals))
            
    f.write(f"Total non-empty rows in first 100 rows: {len(non_empty_rows)}\n")
    for r, vals in non_empty_rows:
        f.write(f"Row {r}: {vals}\n")

print("Done writing inspect_formulas_res.txt")
