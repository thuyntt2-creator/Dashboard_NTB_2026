import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "search_excel_content_res.txt")

targets = [371218, 39567, 65687, 51274, 35931]
target_strs = [str(t) for t in targets]

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            f.write(f"Searching sheet {name}...\n")
            ws = wb[name]
            row_idx = 1
            for row in ws.iter_rows(values_only=True):
                for col_idx, val in enumerate(row):
                    if val is not None:
                        val_str = str(val).strip().replace(",", "").replace(".", "")
                        for t_str in target_strs:
                            if t_str in val_str:
                                f.write(f"FOUND: sheet={name}, row={row_idx}, col={col_idx+1}, val={val}\n")
                row_idx += 1
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done searching Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx.")
