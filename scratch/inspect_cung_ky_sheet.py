import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_cung_ky_sheet_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        # Load with read_only=True and data_only=True
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        f.write(f"Workbook sheets: {wb.sheetnames}\n")
        
        sheet_name = None
        for s in wb.sheetnames:
            if "cùng kỳ" in s.lower() or "cung ky" in s.lower():
                sheet_name = s
                break
                
        f.write(f"Matched sheet: {sheet_name}\n")
        if sheet_name:
            ws = wb[sheet_name]
            f.write("Reading first 100 rows...\n")
            row_idx = 1
            for row in ws.iter_rows(max_row=100, max_col=15, values_only=True):
                if any(v is not None for v in row):
                    # convert row values to strings
                    row_str = [str(v) if v is not None else "" for v in row]
                    f.write(f"Row {row_idx}: {row_str}\n")
                row_idx += 1
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done fast inspecting Thứ cùng kỳ sheet.")
