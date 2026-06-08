import os
import openpyxl
import pandas as pd

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
targets = [
    371218, 39567, 65687, 51274, 35931,
    44558, 4764, 7409, 6269, 4738, # Đắk Nông
    36518, 3725, 6249, 4688, 3715, # Ninh Thuận
    88611, 9190, 14658, 11914, 9243, # Lâm Đồng
    104719, 11517, 19228, 14425, 9564, # Khánh Hòa
    96812, 10371, 18143, 13978, 8671 # Bình Thuận
]

# Let's normalize target formats
target_strs = set()
for t in targets:
    target_strs.add(str(t))
    target_strs.add(f"{t:,}") # 39,567
    target_strs.add(f"{t/1000:.3f}") # 39.567

print(f"Target strings: {target_strs}")

for filename in os.listdir(workspace_dir):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(workspace_dir, filename)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(r, c).value
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str in target_strs or any(str(t) in val_str for t in targets):
                                # check if it matches any target
                                for t in targets:
                                    if str(t) in val_str or f"{t:,}" in val_str:
                                        print(f"Excel {filename} | Sheet {sheetname} | Row {r}, Col {c} | Val={val}")
            wb.close()
        except Exception as e:
            print(f"Error reading {filename}: {e}")
            
    elif filename.endswith(".csv"):
        file_path = os.path.join(workspace_dir, filename)
        try:
            df = pd.read_csv(file_path)
            for c in df.columns:
                for idx, val in enumerate(df[c]):
                    if pd.notna(val):
                        val_str = str(val).strip()
                        for t in targets:
                            if str(t) in val_str or f"{t:,}" in val_str:
                                print(f"CSV {filename} | Col {c} | Row {idx+2} | Val={val}")
        except Exception as e:
            print(f"Error reading {filename}: {e}")
