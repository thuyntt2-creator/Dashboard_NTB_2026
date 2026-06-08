import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "buu_cuc_bat_on.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_bat_on_cocau_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Cocau" in wb.sheetnames:
            ws = wb["Cocau"]
            f.write("Cocau sheet exists. First 30 rows:\n")
            row_idx = 1
            for row in ws.iter_rows(max_row=30, max_col=15, values_only=True):
                if any(v is not None for v in row):
                    row_str = [str(v) if v is not None else "" for v in row]
                    f.write(f"{row_str}\n")
                row_idx += 1
        else:
            f.write("Cocau sheet does not exist in buu_cuc_bat_on.xlsx\n")
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting buu_cuc_bat_on.xlsx Cocau.")
