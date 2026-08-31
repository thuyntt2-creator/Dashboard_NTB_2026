import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Desktop\New folder\downloaded_user_sheet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['LTC']

row = 15 # Row 15 is Pham Ba Thanh Cong in Col L
print(f"Row {row}:")
for c in range(1, 23):
    col_letter = openpyxl.utils.get_column_letter(c)
    print(f"  Col {col_letter}: value={ws.cell(row=row, column=c).value}")
