import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("vols_tao_don.xlsx", read_only=True)
sheet = wb.active
print("Active sheet name:", sheet.title)

# Read header
header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
print("Header:", header)

date_idx = header.index("Date") if "Date" in header else -1
bc_idx = header.index("Bưu cục") if "Bưu cục" in header else -1
vol_idx = header.index("Volume") if "Volume" in header else -1

if date_idx == -1 or bc_idx == -1 or vol_idx == -1:
    print("Required columns not found!")
    sys.exit(0)

# Sample some rows from the end
rows = list(sheet.iter_rows(values_only=True))
total_rows = len(rows)
print("Total rows including header:", total_rows)

dates = set()
for r in rows[1:]:
    if r[date_idx]:
        dates.add(str(r[date_idx])[:10])

print("All unique dates in Excel:", sorted(list(dates)))
