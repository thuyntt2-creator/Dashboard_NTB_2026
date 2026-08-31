import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
user_file = os.path.join(workspace_dir, "downloaded_user_sheet.xlsx")
output_file = os.path.join(workspace_dir, "scratch", "inspect_lc_formulas_res.txt")

with open(output_file, "w", encoding="utf-8") as f:
    wb = openpyxl.load_workbook(user_file, data_only=False)
    if "data rớt LC" in wb.sheetnames:
        ws = wb["data rớt LC"]
        f.write("--- data rớt LC sheet formulas ---\n")
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
        f.write("Row 1 (headers): " + str([ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]) + "\n")
        for r in range(2, 15):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            f.write(f"Row {r}: {row_vals}\n")
            
        # Let's inspect some rows with non-zero %_rot_lc (column 5)
        # We can open the data-only workbook too to find rows where %_rot_lc is not '0.0%'
        wb_data = openpyxl.load_workbook(user_file, data_only=True)
        ws_data = wb_data["data rớt LC"]
        f.write("\nRows where %_rot_lc is non-zero in data-only:\n")
        count = 0
        for r in range(2, ws_data.max_row + 1):
            pct_val = ws_data.cell(row=r, column=5).value
            vol_can = ws_data.cell(row=r, column=4).value
            vol_rot = ws_data.cell(row=r, column=8).value
            tuan = ws_data.cell(row=r, column=6).value
            
            # Print if pct_val is non-zero
            if pct_val and pct_val != "0.0%" and pct_val != 0.0:
                f.write(f"Row {r}: {ws_data.cell(row=r, column=2).value} | Vol cần: {vol_can} | % rớt: {pct_val} | Vol rớt: {vol_rot} | Tuần: {tuan}\n")
                f.write(f"Formulas: {ws.cell(row=r, column=2).value} | {ws.cell(row=r, column=4).value} | {ws.cell(row=r, column=5).value} | {ws.cell(row=r, column=8).value}\n")
                count += 1
                if count >= 10:
                    break
    else:
        f.write("data rớt LC not found in workbook\n")
