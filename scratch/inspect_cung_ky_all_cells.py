import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_cung_ky_all_cells.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = None
        for s in wb.sheetnames:
            if "cùng kỳ" in s.lower() or "cung ky" in s.lower():
                sheet_name = s
                break
                
        f.write(f"Matched sheet: {sheet_name}\n")
        if sheet_name:
            ws = wb[sheet_name]
            non_empty_count = 0
            row_idx = 1
            for row in ws.iter_rows(values_only=True):
                # check if there are any non-empty values in the row
                non_empty_cols = [(col_idx+1, val) for col_idx, val in enumerate(row) if val is not None]
                if non_empty_cols:
                    non_empty_count += 1
                    f.write(f"Row {row_idx}: {non_empty_cols}\n")
                    if non_empty_count > 500:
                        f.write("Truncated after 500 non-empty rows.\n")
                        break
                row_idx += 1
            f.write(f"Total non-empty rows: {non_empty_count}\n")
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting all cells.")
