import openpyxl
import pandas as pd
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Desktop\New folder\NTB_Bao_Cao_Van_Hanh_Corrected.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

if 'Phân tích AM W24 vs W23' in wb.sheetnames:
    ws = wb['Phân tích AM W24 vs W23']
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    # Let's find Nguyễn Duy Long row
    headers = data[0]
    df = pd.DataFrame(data[1:], columns=headers)
    # Filter out empty rows or rows where first column is None
    df = df[df.iloc[:, 0].notna()]
    
    ndl_rows = df[df.iloc[:, 0] == 'Nguyễn Duy Long']
    print("Local Excel Nguyễn Duy Long rows:")
    for idx, row in ndl_rows.iterrows():
        print(f"Row {idx}:")
        for col in df.columns:
            print(f"  {col}: {row[col]}")
else:
    print("Worksheet 'Phân tích AM W24 vs W23' NOT found in local Excel!")
