import pandas as pd
import openpyxl

path = "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx"
wb = openpyxl.load_workbook(path, read_only=True)
print("Sheet names:", wb.sheetnames)

# Inspect first few rows of "Bưu cục"
ws_bc = wb["Bưu cục"]
print("\n--- Bưu cục sheet header (first 5 rows) ---")
for r in range(1, 15):
    row_vals = [ws_bc.cell(row=r, column=c).value for c in range(1, 25)]
    if any(row_vals):
        print(f"Row {r}: {row_vals}")

# Inspect first few rows of "Ca1 - Ca2 - Tồn"
ws_ca = wb["Ca1 - Ca2 - Tồn"]
print("\n--- Ca1 - Ca2 - Tồn sheet header (first 5 rows) ---")
for r in range(1, 15):
    row_vals = [ws_ca.cell(row=r, column=c).value for c in range(1, 25)]
    if any(row_vals):
        print(f"Row {r}: {row_vals}")
