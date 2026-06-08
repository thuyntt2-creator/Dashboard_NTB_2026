import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "OPR TTS.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_opr_raw_sheets_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheetname in ['rawopr', 'raw', 'REPORT_OPR', 'CoCauVung']:
            if sheetname in wb.sheetnames:
                ws = wb[sheetname]
                f.write(f"\nSheet {sheetname} exists. First 10 rows:\n")
                row_idx = 1
                for row in ws.iter_rows(max_row=10, max_col=15, values_only=True):
                    if any(v is not None for v in row):
                        row_str = [str(v) if v is not None else "" for v in row]
                        f.write(f"{row_str}\n")
                    row_idx += 1
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting OPR raw sheets.")
