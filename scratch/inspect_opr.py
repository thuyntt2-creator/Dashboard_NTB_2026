import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "OPR TTS.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_opr_sheet_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if 'OPR' in wb.sheetnames:
            ws = wb['OPR']
            f.write("OPR sheet exists. Reading first 50 rows...\n")
            for row in ws.iter_rows(max_row=50, max_col=15, values_only=True):
                if any(v is not None for v in row):
                    row_str = [str(v) if v is not None else "" for v in row]
                    f.write(f"{row_str}\n")
        else:
            f.write("OPR sheet does not exist.\n")
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting OPR sheet.")
