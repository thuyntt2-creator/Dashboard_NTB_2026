import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_AM_Project_3_fixed_5\BaoCao_AM_Project\output\BaoCao_Tuan_NTB_W35_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['01_San luong']

# 1. AM Full hàng (rows 12-29)
am_full = []
for r in range(12, 30):
    am = ws.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM', 'THEO AM - TTS']:
        w32 = ws.cell(r, 2).value or 0
        w33 = ws.cell(r, 3).value or 0
        w34 = ws.cell(r, 4).value or 0
        w35 = ws.cell(r, 5).value or 0
        diff = ws.cell(r, 6).value if ws.cell(r, 6).value is not None else (w35 - w34)
        am_full.append({
            'am': str(am).strip(),
            'vol': w35,
            'w32': w32,
            'w33': w33,
            'w34': w34,
            'w35': w35,
            'diff': diff
        })

# 2. AM TTS (rows 34-51)
am_tts = []
for r in range(34, 52):
    am = ws.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM', 'THEO TỈNH - Full hàng']:
        w32 = ws.cell(r, 2).value or 0
        w33 = ws.cell(r, 3).value or 0
        w34 = ws.cell(r, 4).value or 0
        w35 = ws.cell(r, 5).value or 0
        diff = ws.cell(r, 6).value if ws.cell(r, 6).value is not None else (w35 - w34)
        am_tts.append({
            'am': str(am).strip(),
            'vol': w35,
            'w32': w32,
            'w33': w33,
            'w34': w34,
            'w35': w35,
            'diff': diff
        })

# 3. Tỉnh Full hàng (rows 56-60)
tinh_full = []
for r in range(56, 61):
    t = ws.cell(r, 1).value
    if t and str(t).strip() not in ['None', '', 'Tỉnh', 'THEO TỈNH - TTS']:
        w32 = ws.cell(r, 2).value or 0
        w33 = ws.cell(r, 3).value or 0
        w34 = ws.cell(r, 4).value or 0
        w35 = ws.cell(r, 5).value or 0
        diff = ws.cell(r, 6).value if ws.cell(r, 6).value is not None else (w35 - w34)
        tinh_full.append({
            'tinh': str(t).strip(),
            'vol': w35,
            'w32': w32,
            'w33': w33,
            'w34': w34,
            'w35': w35,
            'diff': diff
        })

# 4. Tỉnh TTS (rows 65-69)
tinh_tts = []
for r in range(65, 70):
    t = ws.cell(r, 1).value
    if t and str(t).strip() not in ['None', '', 'Tỉnh']:
        w32 = ws.cell(r, 2).value or 0
        w33 = ws.cell(r, 3).value or 0
        w34 = ws.cell(r, 4).value or 0
        w35 = ws.cell(r, 5).value or 0
        diff = ws.cell(r, 6).value if ws.cell(r, 6).value is not None else (w35 - w34)
        tinh_tts.append({
            'tinh': str(t).strip(),
            'vol': w35,
            'w32': w32,
            'w33': w33,
            'w34': w34,
            'w35': w35,
            'diff': diff
        })

print("Parsed AM FULL Count:", len(am_full))
print("Sample AM FULL:", am_full[:2])
print("\nParsed AM TTS Count:", len(am_tts))
print("Sample AM TTS:", am_tts[:2])
print("\nParsed TINH FULL Count:", len(tinh_full))
print("Sample TINH FULL:", tinh_full[:2])
print("\nParsed TINH TTS Count:", len(tinh_tts))
print("Sample TINH TTS:", tinh_tts[:2])
