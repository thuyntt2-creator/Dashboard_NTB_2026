import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "temp_ops_check.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_temp_ops_cung_ky_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Thứ cùng kỳ" in wb.sheetnames:
            ws = wb["Thứ cùng kỳ"]
            f.write("Thứ cùng kỳ exists in temp_ops_check.xlsx\n")
            non_empty_count = 0
            row_idx = 1
            for row in ws.iter_rows(values_only=True):
                non_empty_cols = [(col_idx+1, val) for col_idx, val in enumerate(row) if val is not None]
                if non_empty_cols:
                    non_empty_count += 1
                    f.write(f"Row {row_idx}: {non_empty_cols}\n")
                    if non_empty_count > 100:
                        f.write("Truncated.\n")
                        break
                row_idx += 1
            f.write(f"Total non-empty rows: {non_empty_count}\n")
        else:
            f.write("Thứ cùng kỳ does not exist in temp_ops_check.xlsx\n")
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting temp_ops_check.xlsx.")
