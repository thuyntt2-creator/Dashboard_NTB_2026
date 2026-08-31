import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Desktop\New folder\downloaded_user_sheet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['LTC']

row_idx_A = None
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val == 'Phạm Bá Thành Công':
        row_idx_A = r
        break

row_idx_M = None
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=13).value # Column M is 13
    if val == 'Phạm Bá Thành Công':
        row_idx_M = r
        break

print(f"Phạm Bá Thành Công in Col A: row {row_idx_A}")
print(f"Phạm Bá Thành Công in Col M: row {row_idx_M}")

if row_idx_A:
    print(f"\nRow {row_idx_A} (from Col A):")
    for c in range(1, 23):
        col_letter = openpyxl.utils.get_column_letter(c)
        print(f"  Col {col_letter}: {ws.cell(row=row_idx_A, column=c).value}")

if row_idx_M and row_idx_M != row_idx_A:
    print(f"\nRow {row_idx_M} (from Col M):")
    for c in range(1, 23):
        col_letter = openpyxl.utils.get_column_letter(c)
        print(f"  Col {col_letter}: {ws.cell(row=row_idx_M, column=c).value}")
