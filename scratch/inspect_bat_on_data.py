import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "buu_cuc_bat_on.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_bat_on_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        f.write(f"Workbook sheets: {wb.sheetnames}\n")
        
        for sheetname in ['Data', 'NTB']:
            if sheetname in wb.sheetnames:
                ws = wb[sheetname]
                f.write(f"\nSheet: {sheetname}\n")
                row_idx = 1
                for row in ws.iter_rows(max_row=10, max_col=15, values_only=True):
                    if any(v is not None for v in row):
                        row_str = [str(v) if v is not None else "" for v in row]
                        f.write(f"Row {row_idx}: {row_str}\n")
                    row_idx += 1
        wb.close()
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Done inspecting buu_cuc_bat_on.xlsx sheets.")
