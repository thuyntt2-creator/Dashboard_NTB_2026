import pandas as pd
import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"

def main():
    print("Searching excel files for customer group revenue numbers...")
    for filename in os.listdir(workspace_dir):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            file_path = os.path.join(workspace_dir, filename)
            print(f"Checking {filename}...")
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for r_idx in range(1, min(ws.max_row + 1, 1000)):
                        for c_idx in range(1, min(ws.max_column + 1, 50)):
                            val = ws.cell(row=r_idx, column=c_idx).value
                            if val is not None:
                                val_str = str(val)
                                if "so tuần trước" in val_str or "309,8" in val_str or "251,7" in val_str:
                                    print(f"  FOUND in {filename} | Sheet: {sheet} | Cell: {openpyxl.utils.get_column_letter(c_idx)}{r_idx} | Value: {val_str}")
            except Exception as e:
                print(f"  Error checking {filename}: {e}")

if __name__ == '__main__':
    main()
