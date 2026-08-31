import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Desktop\New folder\downloaded_user_sheet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['gtcnew']

print("Formulas in gtcnew sheet row 31 to 36:")
for r in range(31, 37):
    print(f"Row {r}:")
    print(f"  Col A: {ws.cell(row=r, column=1).value}")
    print(f"  Col B: {ws.cell(row=r, column=2).value}")
    print(f"  Col C: {ws.cell(row=r, column=3).value}")
