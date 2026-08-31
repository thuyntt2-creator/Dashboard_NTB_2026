import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Desktop\New folder\NTB_Bao_Cao_Van_Hanh_Corrected.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

if 'gtcnew' in wb.sheetnames:
    ws = wb['gtcnew']
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=13).value # Column M is 13
        if val == 'Nguyễn Duy Long':
            print(f"Local Excel row {r} for Nguyễn Duy Long in gtcnew tab (TTS table):")
            # Print columns M to U (13 to 21)
            for c in range(13, 22):
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"  Col {col_letter}: {ws.cell(row=r, column=c).value}")
else:
    print("Sheet 'gtcnew' not found in local Excel!")
