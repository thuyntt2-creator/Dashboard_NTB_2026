import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_thu_cung_ky_openpyxl.txt")

os.makedirs(os.path.dirname(output_path), exist_ok=True)

with open(output_path, "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = None
        for s in wb.sheetnames:
            if "cùng kỳ" in s.lower() or "cung ky" in s.lower():
                sheet_name = s
                break
        f.write(f"Matched sheet: {sheet_name}\n")
        if sheet_name:
            ws = wb[sheet_name]
            f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
            for r in range(1, min(100, ws.max_row + 1)):
                row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
                row_str = [str(v) if v is not None else "" for v in row_vals]
                if any(row_str):
                    f.write(f"Row {r}: {row_str}\n")
        else:
            f.write("No sheet matched.\n")
    except Exception as e:
        f.write(f"Error: {e}\n")

print("Inspection completed.")
