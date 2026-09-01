import openpyxl, sys

sys.stdout.reconfigure(encoding='utf-8')
excel_file = r'C:\Users\lap4all\Downloads\BaoCao_AM_Project_3_fixed_5\BaoCao_AM_Project\output\BaoCao_Tuan_NTB_W35_2026.xlsx'
wb = openpyxl.load_workbook(excel_file, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f'==================================================')
    print(f'SHEET: {name} (max_row={ws.max_row}, max_col={ws.max_column})')
    print(f'==================================================')
    for r in range(1, ws.max_row+1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        if c1 is not None and str(c1).strip() != '':
            # check if header
            val_str = str(c1).strip()
            if any(k in val_str.upper() for k in ['THEO', 'TỔNG', 'CHỈ TIÊU', 'BẢNG', 'AM', 'TỈNH', 'KHUNG GIỜ', 'NGUỒN', 'CHỈ SỐ', 'PHÂN KHÚC']):
                print(f'  [SECTION/HEADER R{r:02d}]: {val_str} | col2={c2} | col3={c3}')
            elif r in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 30, 31, 32, 33, 34, 52, 53, 54, 55, 56, 61, 62, 63, 64, 65]:
                print(f'  [DATA SAMPLE R{r:02d}]: c1={c1} | c2={c2} | c3={c3}')
