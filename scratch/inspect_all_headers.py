import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_AM_Project_3_fixed_5\BaoCao_AM_Project\output\BaoCao_Tuan_NTB_W35_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== SHEET: {name} ===")
    for r in range(1, min(15, ws.max_row+1)):
        c1 = ws.cell(r, 1).value
        if c1 and any(k in str(c1).upper() for k in ['AM', 'CHỈ TIÊU', 'TỈNH', 'STT']):
            cols = [ws.cell(r, c).value for c in range(1, ws.max_column+1) if ws.cell(r, c).value is not None]
            print(f"  Header R{r:02d} ({len(cols)} cols): {cols}")
    # print sample row 12
    if ws.max_row >= 12:
        row12 = [ws.cell(12, c).value for c in range(1, 10)]
        print(f"  Sample R12: {row12}")
