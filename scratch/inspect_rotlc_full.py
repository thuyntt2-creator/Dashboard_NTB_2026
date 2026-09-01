import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_AM_Project_3_fixed_5\BaoCao_AM_Project\output\BaoCao_Tuan_NTB_W35_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['09_Rot LC']

print("=== SHEET 09_Rot LC ALL ROWS ===")
for r in range(1, ws.max_row+1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in vals):
        print(f"R{r:02d}: {[v for v in vals if v is not None][:8]}")
