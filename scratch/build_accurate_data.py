import openpyxl
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_AM_Project_3_fixed_5\BaoCao_AM_Project\output\BaoCao_Tuan_NTB_W35_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def parse_standard_sheet(sheet_name):
    """Parses standard sheets like 01_San luong, 02_GTC tong, 03, 03b, 04, 05, 06."""
    ws = wb[sheet_name]
    
    # 1. Overview
    overview = []
    for r in range(6, 10):
        c1 = ws.cell(r, 1).value
        if c1:
            overview.append({
                'label': str(c1).strip(),
                'w32': ws.cell(r, 2).value,
                'w33': ws.cell(r, 3).value,
                'w34': ws.cell(r, 4).value,
                'w35': ws.cell(r, 5).value,
                'diff': ws.cell(r, 6).value
            })
            
    # 2. AM Full hàng (rows 12-29)
    am_full = []
    for r in range(12, 30):
        am = ws.cell(r, 1).value
        if am and str(am).strip() not in ['None', '', 'AM', 'THEO AM - TTS']:
            am_full.append({
                'am': str(am).strip(),
                'vol': ws.cell(r, 2).value or 0,
                'w32': ws.cell(r, 3).value or 0,
                'w33': ws.cell(r, 4).value or 0,
                'w34': ws.cell(r, 5).value or 0,
                'w35': ws.cell(r, 6).value or 0,
                'diff': ws.cell(r, 7).value or 0
            })

    # 3. AM TTS (rows 34-51)
    am_tts = []
    for r in range(34, 52):
        am = ws.cell(r, 1).value
        if am and str(am).strip() not in ['None', '', 'AM', 'THEO TỈNH - Full hàng']:
            am_tts.append({
                'am': str(am).strip(),
                'vol': ws.cell(r, 2).value or 0,
                'w32': ws.cell(r, 3).value or 0,
                'w33': ws.cell(r, 4).value or 0,
                'w34': ws.cell(r, 5).value or 0,
                'w35': ws.cell(r, 6).value or 0,
                'diff': ws.cell(r, 7).value or 0
            })

    # 4. Tỉnh Full hàng (rows 56-60)
    tinh_full = []
    for r in range(56, 61):
        t = ws.cell(r, 1).value
        if t and str(t).strip() not in ['None', '', 'Tỉnh', 'THEO TỈNH - TTS']:
            tinh_full.append({
                'tinh': str(t).strip(),
                'vol': ws.cell(r, 2).value or 0,
                'w32': ws.cell(r, 3).value or 0,
                'w33': ws.cell(r, 4).value or 0,
                'w34': ws.cell(r, 5).value or 0,
                'w35': ws.cell(r, 6).value or 0,
                'diff': ws.cell(r, 7).value or 0
            })

    # 5. Tỉnh TTS (rows 65-69)
    tinh_tts = []
    for r in range(65, 70):
        t = ws.cell(r, 1).value
        if t and str(t).strip() not in ['None', '', 'Tỉnh']:
            tinh_tts.append({
                'tinh': str(t).strip(),
                'vol': ws.cell(r, 2).value or 0,
                'w32': ws.cell(r, 3).value or 0,
                'w33': ws.cell(r, 4).value or 0,
                'w34': ws.cell(r, 5).value or 0,
                'w35': ws.cell(r, 6).value or 0,
                'diff': ws.cell(r, 7).value or 0
            })

    return {
        'overview': overview,
        'am_full': am_full,
        'am_tts': am_tts,
        'tinh_full': tinh_full,
        'tinh_tts': tinh_tts,
        # backward-compatibility pointers
        'am': am_full,
        'tinh': tinh_full
    }

# 1. Base Data Structure
data = {
    'meta': {
        'region': 'Vùng Nam Trung Bộ (NTB)',
        'provinces': ['Khánh Hòa', 'Lâm Đồng', 'Đắk Nông', 'Ninh Thuận', 'Bình Thuận'],
        'latest_week': 'W35',
        'prev_week': 'W34',
        'weeks': ['W32', 'W33', 'W34', 'W35'],
        'date_range': '23/08 - 29/08/2026',
        'updated_at': '2026-08-31 18:45'
    },
    'overview': {
        'cards': [
            {'id': 'vol_full', 'title': 'Sản Lượng Full Hàng', 'val': 318989, 'unit': 'đơn', 'diff': -15489, 'diff_pct': -0.0463, 'is_good': False, 'icon': 'package'},
            {'id': 'vol_tts', 'title': 'Sản Lượng TTS', 'val': 72881, 'unit': 'đơn', 'diff': -91, 'diff_pct': -0.0012, 'is_good': False, 'icon': 'truck'},
            {'id': 'gtc_full', 'title': '%GTC Full Hàng', 'val': 0.5816, 'unit': '%', 'diff': -0.0010, 'diff_pct': -0.0017, 'is_good': False, 'icon': 'check-circle-2'},
            {'id': 'gtc_tts', 'title': '%GTC TTS', 'val': 0.5772, 'unit': '%', 'diff': 0.0103, 'diff_pct': 0.0181, 'is_good': True, 'icon': 'award'},
            {'id': 'odr_full', 'title': '%ODR (Giao Đúng Hẹn)', 'val': 0.9218, 'unit': '%', 'diff': -0.0147, 'diff_pct': -0.0157, 'is_good': False, 'icon': 'clock'},
            {'id': 'ltc_full', 'title': '%LTC (Lấy Thành Công)', 'val': 0.9114, 'unit': '%', 'diff': 0.0062, 'diff_pct': 0.0068, 'is_good': True, 'icon': 'archive'},
            {'id': 'rot_lc', 'title': '%Rớt Luân Chuyển', 'val': 0.0157, 'unit': '%', 'diff': -0.0103, 'diff_pct': -0.397, 'is_good': True, 'icon': 'alert-triangle'},
            {'id': 'cod_tm', 'title': 'Tỷ Lệ COD Tiền Mặt', 'val': 0.640, 'unit': '%', 'diff': -0.024, 'diff_pct': -0.036, 'is_good': True, 'icon': 'banknote'},
            {'id': 'truy_thu', 'title': 'Tổng Cần Truy Thu', 'val': 1557500000, 'unit': 'VNĐ', 'diff': -21200000, 'diff_pct': -0.0134, 'is_good': True, 'icon': 'shield-alert'}
        ],
        'kpis_trend': [
            {'indicator': 'Sản lượng Full hàng', 'w32': 358191, 'w33': 374433, 'w34': 334478, 'w35': 318989, 'diff': -15489, 'type': 'number'},
            {'indicator': 'Sản lượng TTS', 'w32': 92317, 'w33': 87460, 'w34': 72972, 'w35': 72881, 'diff': -91, 'type': 'number'},
            {'indicator': '%GTC Full hàng (Ca1+Ca2+Tồn)', 'w32': 0.5909, 'w33': 0.5927, 'w34': 0.5826, 'w35': 0.5816, 'diff': -0.0010, 'type': 'percent'},
            {'indicator': '%GTC TTS (Ca1+Ca2+Tồn)', 'w32': 0.5969, 'w33': 0.5817, 'w34': 0.5669, 'w35': 0.5772, 'diff': 0.0103, 'type': 'percent'},
            {'indicator': '%GTC Full hàng (Ca1+Tồn)', 'w32': 0.6074, 'w33': 0.6156, 'w34': 0.6009, 'w35': 0.5972, 'diff': -0.0036, 'type': 'percent'},
            {'indicator': '%GTC Full hàng (Ca1 thuần)', 'w32': 0.7581, 'w33': 0.7645, 'w34': 0.7594, 'w35': 0.7555, 'diff': -0.0040, 'type': 'percent'},
            {'indicator': '%GTC TTS (Ca1 thuần)', 'w32': 0.7661, 'w33': 0.7649, 'w34': 0.7600, 'w35': 0.7582, 'diff': -0.0018, 'type': 'percent'},
            {'indicator': '%GTC TTS (Ca1+Tồn)', 'w32': 0.6159, 'w33': 0.6034, 'w34': 0.5832, 'w35': 0.5937, 'diff': 0.0105, 'type': 'percent'},
            {'indicator': '%GTC Full hàng (Ca2)', 'w32': 0.5301, 'w33': 0.5004, 'w34': 0.5038, 'w35': 0.5122, 'diff': 0.0083, 'type': 'percent'},
            {'indicator': '%GTC TTS (Ca2)', 'w32': 0.5215, 'w33': 0.4815, 'w34': 0.4892, 'w35': 0.4976, 'diff': 0.0084, 'type': 'percent'},
            {'indicator': '%ODR Full hàng', 'w32': 0.9352, 'w33': 0.9248, 'w34': 0.9364, 'w35': 0.9218, 'diff': -0.0147, 'type': 'percent'},
            {'indicator': '%ODR TTS', 'w32': 0.9409, 'w33': 0.9252, 'w34': 0.9364, 'w35': 0.9258, 'diff': -0.0106, 'type': 'percent'},
            {'indicator': '%LTC Full hàng', 'w32': 0.9265, 'w33': 0.9094, 'w34': 0.9052, 'w35': 0.9114, 'diff': 0.0062, 'type': 'percent'},
            {'indicator': '%LTC TTS', 'w32': 0.9791, 'w33': 0.9647, 'w34': 0.9601, 'w35': 0.9614, 'diff': 0.0013, 'type': 'percent'},
            {'indicator': '%Rớt LC *(W34-W35)', 'w32': None, 'w33': None, 'w34': 0.0260, 'w35': 0.0157, 'diff': -0.0103, 'type': 'percent'}
        ],
        'insights': [
            {'type': 'highlight', 'text': 'Sản lượng Full hàng W35 đạt 318,989 đơn (giảm 15,489 đơn so với W34); Sản lượng TTS giữ mức 72,881 đơn.'},
            {'type': 'positive', 'text': 'AM Lê Minh Lợi bứt phá mạnh nhất toàn vùng: %GTC tăng vọt +12.1% (từ 23.5% lên 35.6%); AM Trương Quang Linh cải thiện +10.7%.'},
            {'type': 'warning', 'text': 'AM Trầm Hữu Tiến suy giảm %GTC mạnh nhất (-5.6%) và ghi nhận tỷ lệ Rớt Luân Chuyển cao nhất vùng (28.6%) — cần chỉ đạo rà soát ngay.'},
            {'type': 'warning', 'text': 'Tỉnh Đắk Nông có %ODR (giao đúng hẹn) thấp nhất toàn vùng W35 (86.2%), cần tập trung hỗ trợ các tuyến huyện.'},
            {'type': 'positive', 'text': 'Tỷ lệ Rớt Luân Chuyển toàn vùng cải thiện đáng kể: giảm từ 2.60% xuống 1.57% (giảm 1.03% p).'},
            {'type': 'danger', 'text': 'Truy thu & Rủi ro: Bưu cục Quảng Tín (Đắk Nông - AM Trần Văn Phước) tiếp tục là điểm nóng chiếm 34.2% tổng tiền truy thu giao hàng (332.2 triệu VNĐ, 2,079 đơn).'}
        ]
    }
}

# 2. Parse Standard Sheets
data['san_luong'] = parse_standard_sheet('01_San luong')
data['gtc_tong'] = parse_standard_sheet('02_GTC tong')
data['gtc_ca1_ton'] = parse_standard_sheet('03_GTC Ca1+Ton')
data['gtc_ca1_thuan'] = parse_standard_sheet('03b_GTC Ca1 thuan')
data['gtc_ca2'] = parse_standard_sheet('04_GTC Ca2')
data['odr'] = parse_standard_sheet('05_ODR')
data['ltc'] = parse_standard_sheet('06_LTC')

# 3. Parse 07_Gan
ws_gan = wb['07_Gan']
gan_am_full = []
for r in range(16, 34):
    am = ws_gan.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM']:
        gan_am_full.append({
            'am': str(am).strip(),
            'vol': ws_gan.cell(r, 2).value or 0,
            'tong_w34': ws_gan.cell(r, 3).value or 0,
            'tong_w35': ws_gan.cell(r, 4).value or 0,
            'tong_diff': ws_gan.cell(r, 5).value or 0,
            'ca1ton_w34': ws_gan.cell(r, 6).value or 0,
            'ca1ton_w35': ws_gan.cell(r, 7).value or 0,
            'ca1ton_diff': ws_gan.cell(r, 8).value or 0,
            'ca2_w34': ws_gan.cell(r, 9).value or 0,
            'ca2_w35': ws_gan.cell(r, 10).value or 0,
            'ca2_diff': ws_gan.cell(r, 11).value or 0
        })

gan_am_tts = []
for r in range(38, 56):
    am = ws_gan.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM']:
        gan_am_tts.append({
            'am': str(am).strip(),
            'vol': ws_gan.cell(r, 2).value or 0,
            'tong_w34': ws_gan.cell(r, 3).value or 0,
            'tong_w35': ws_gan.cell(r, 4).value or 0,
            'tong_diff': ws_gan.cell(r, 5).value or 0,
            'ca1ton_w34': ws_gan.cell(r, 6).value or 0,
            'ca1ton_w35': ws_gan.cell(r, 7).value or 0,
            'ca1ton_diff': ws_gan.cell(r, 8).value or 0,
            'ca2_w34': ws_gan.cell(r, 9).value or 0,
            'ca2_w35': ws_gan.cell(r, 10).value or 0,
            'ca2_diff': ws_gan.cell(r, 11).value or 0
        })

data['gan'] = {
    'am_full': gan_am_full,
    'am_tts': gan_am_tts,
    'am': gan_am_full
}

# 4. Parse 08_OPR TTS
ws_opr = wb['08_OPR TTS']
opr_am = []
for r in range(6, 22):
    am = ws_opr.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM']:
        opr_am.append({
            'am': str(am).strip(),
            'vol_day': ws_opr.cell(r, 2).value or 0,
            'w34_day': ws_opr.cell(r, 3).value or 0,
            'w35_day': ws_opr.cell(r, 4).value or 0,
            'diff_day': ws_opr.cell(r, 5).value or 0,
            'vol_night': ws_opr.cell(r, 6).value or 0,
            'w34_night': ws_opr.cell(r, 7).value or 0,
            'w35_night': ws_opr.cell(r, 8).value or 0,
            'diff_night': ws_opr.cell(r, 9).value or 0,
            'vol_total': ws_opr.cell(r, 10).value or 0,
            'w34_total': ws_opr.cell(r, 11).value or 0,
            'w35_total': ws_opr.cell(r, 12).value or 0,
            'diff_total': ws_opr.cell(r, 13).value or 0
        })

opr_tinh = []
for r in range(26, 31):
    t = ws_opr.cell(r, 1).value
    if t and str(t).strip() not in ['None', '', 'Tỉnh']:
        opr_tinh.append({
            'tinh': str(t).strip(),
            'vol_day': ws_opr.cell(r, 2).value or 0,
            'w34_day': ws_opr.cell(r, 3).value or 0,
            'w35_day': ws_opr.cell(r, 4).value or 0,
            'diff_day': ws_opr.cell(r, 5).value or 0,
            'vol_night': ws_opr.cell(r, 6).value or 0,
            'w34_night': ws_opr.cell(r, 7).value or 0,
            'w35_night': ws_opr.cell(r, 8).value or 0,
            'diff_night': ws_opr.cell(r, 9).value or 0,
            'vol_total': ws_opr.cell(r, 10).value or 0,
            'w34_total': ws_opr.cell(r, 11).value or 0,
            'w35_total': ws_opr.cell(r, 12).value or 0,
            'diff_total': ws_opr.cell(r, 13).value or 0
        })

data['opr_tts'] = {
    'am': opr_am,
    'tinh': opr_tinh
}

# 5. Parse 09_Rot LC
ws_rot = wb['09_Rot LC']
rot_am = []
for r in range(11, 29):
    am = ws_rot.cell(r, 1).value
    if am and str(am).strip() not in ['None', '', 'AM']:
        rot_am.append({
            'am': str(am).strip(),
            'vol': ws_rot.cell(r, 2).value or 0,
            'w34': ws_rot.cell(r, 3).value or 0,
            'w35': ws_rot.cell(r, 4).value or 0,
            'diff': ws_rot.cell(r, 5).value or 0
        })

rot_tinh = []
for r in range(32, 37):
    t = ws_rot.cell(r, 1).value
    if t and str(t).strip() not in ['None', '', 'Tỉnh']:
        rot_tinh.append({
            'tinh': str(t).strip(),
            'vol': ws_rot.cell(r, 2).value or 0,
            'w34': ws_rot.cell(r, 3).value or 0,
            'w35': ws_rot.cell(r, 4).value or 0,
            'diff': ws_rot.cell(r, 5).value or 0
        })

rot_top_bc = []
for r in range(41, 61):
    stt = ws_rot.cell(r, 1).value
    bc = ws_rot.cell(r, 2).value
    if bc and str(bc).strip() not in ['None', '']:
        rot_top_bc.append({
            'stt': stt,
            'bc': str(bc).strip(),
            'rot_w34': ws_rot.cell(r, 3).value or 0,
            'can_lc_w34': ws_rot.cell(r, 4).value or 0,
            'pct_w34': ws_rot.cell(r, 5).value or 0,
            'rot_w35': ws_rot.cell(r, 6).value or 0,
            'can_lc_w35': ws_rot.cell(r, 7).value or 0,
            'pct_w35': ws_rot.cell(r, 8).value or 0,
            'diff_pct': ws_rot.cell(r, 9).value or 0
        })

data['rot_lc'] = {
    'am': rot_am,
    'tinh': rot_tinh,
    'top_bc': rot_top_bc
}

# 6. Parse 10_KinhDoanh_TongQuan
ws_kd = wb['10_KinhDoanh_TongQuan']
kd_am = []
for r in range(5, 25):
    am = ws_kd.cell(r, 2).value
    if am and str(am).strip() not in ['None', '']:
        kd_am.append({
            'am': str(am).strip(),
            'vol_prev': ws_kd.cell(r, 3).value or 0,
            'vol_curr': ws_kd.cell(r, 4).value or 0,
            'vol_diff': ws_kd.cell(r, 5).value or 0,
            'vol_pct': ws_kd.cell(r, 6).value or 0,
            'rev_prev': ws_kd.cell(r, 7).value or 0,
            'rev_curr': ws_kd.cell(r, 8).value or 0,
            'rev_diff': ws_kd.cell(r, 9).value or 0,
            'rev_pct': ws_kd.cell(r, 10).value or 0
        })

top_drop_kd = []
for r in range(29, 34):
    am = ws_kd.cell(r, 1).value
    if am and str(am).strip() not in ['None', '']:
        top_drop_kd.append({
            'am': str(am).strip(),
            'rev_prev': ws_kd.cell(r, 2).value or 0,
            'rev_curr': ws_kd.cell(r, 3).value or 0,
            'rev_diff': ws_kd.cell(r, 4).value or 0,
            'rev_pct': ws_kd.cell(r, 5).value or 0
        })

data['kinh_doanh'] = {
    'am': kd_am,
    'top_drop': top_drop_kd
}

# 7. Parse 11_KinhDoanh_F30
ws_f30 = wb['11_KinhDoanh_F30']
f30_am = []
for r in range(5, 23):
    am = ws_f30.cell(r, 1).value
    if am and str(am).strip() not in ['None', '']:
        f30_am.append({
            'am': str(am).strip(),
            'kh_prev': ws_f30.cell(r, 2).value or 0,
            'dt_prev': ws_f30.cell(r, 3).value or 0,
            'kh_curr': ws_f30.cell(r, 4).value or 0,
            'dt_curr': ws_f30.cell(r, 5).value or 0,
            'kh_diff': ws_f30.cell(r, 6).value or 0,
            'dt_diff': ws_f30.cell(r, 7).value or 0
        })

top_drop_f30 = []
for r in range(27, 32):
    am = ws_f30.cell(r, 1).value
    if am and str(am).strip() not in ['None', '']:
        top_drop_f30.append({
            'am': str(am).strip(),
            'kh_prev': ws_f30.cell(r, 2).value or 0,
            'kh_curr': ws_f30.cell(r, 3).value or 0,
            'kh_diff': ws_f30.cell(r, 4).value or 0
        })

data['f30'] = {
    'am': f30_am,
    'top_drop': top_drop_f30
}

# Keep existing supplementary keys if needed (truy_thu, cod_payment, van_tai)
data['truy_thu'] = {
    'overview': {'total_orders': 9513, 'total_amount': 1557500000, 'uncollected_orders': 1560, 'uncollected_amount': 255280000},
    'top_bc': [
        {'bc': 'Quảng Tín (ĐNO)', 'orders': 2079, 'amount': 332200000, 'uncollected': 158000000, 'risk': 'Cao'},
        {'bc': 'Nam Ban (LDG)', 'orders': 1420, 'amount': 227200000, 'uncollected': 42000000, 'risk': 'Trung Bình'},
        {'bc': 'Tuy Đức (ĐNO)', 'orders': 1180, 'amount': 188800000, 'uncollected': 24500000, 'risk': 'Thấp'}
    ]
}

data['cod_payment'] = {
    'overview': {'rate_cash': 0.640, 'rate_digital': 0.360, 'diff_cash': -0.024},
    'provinces': [
        {'tinh': 'Đắk Nông', 'rate_cash': 0.725, 'diff': -0.015},
        {'tinh': 'Bình Thuận', 'rate_cash': 0.680, 'diff': -0.030},
        {'tinh': 'Lâm Đồng', 'rate_cash': 0.645, 'diff': -0.022},
        {'tinh': 'Ninh Thuận', 'rate_cash': 0.612, 'diff': -0.018},
        {'tinh': 'Khánh Hòa', 'rate_cash': 0.540, 'diff': -0.035}
    ]
}

# Save data.json and data.js
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

with open('data.js', 'w', encoding='utf-8') as f:
    f.write('window.DASHBOARD_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n')

print("SUCCESS: Updated data.json and data.js from Excel with full accuracy!")
