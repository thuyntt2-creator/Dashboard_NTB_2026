import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "inspect_ops_workbook_res.txt")

with open(output_path, "w", encoding="utf-8") as f:
    if not os.path.exists(file_path):
        f.write("File not found\n")
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        f.write(f"Sheets: {wb.sheetnames}\n")
        for name in wb.sheetnames:
            sheet = wb[name]
            f.write(f"\n--- Sheet: {name} ---\n")
            f.write(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}\n")
            
            # Print non-empty cells in the first 20 rows and 15 columns
            found_cells = 0
            for r in range(1, min(40, sheet.max_row + 1)):
                row_vals = []
                for c in range(1, min(25, sheet.max_column + 1)):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None:
                        row_vals.append(f"C{c}:{repr(val)}")
                if row_vals:
                    f.write(f"Row {r}: {', '.join(row_vals)}\n")
                    found_cells += len(row_vals)
            f.write(f"Found {found_cells} non-empty cells in the inspected header area.\n")

print("Done writing inspect_ops_workbook_res.txt")
